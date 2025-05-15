from openpyxl import Workbook
import os

wb = Workbook()
wb.create_sheet('общий отчет')

default_sheet = 'Sheet'
if default_sheet in wb.sheetnames:
    sheets_to_remove = wb[default_sheet]
    wb.remove(sheets_to_remove)

folder_path = os.path.abspath(r'C:\Users\Инстаблогерша\OneDrive\Desktop\bot_py\hw_on_differential_equations_bot\bot\data')

if not os.path.exists(folder_path):
    try:
        os.makedirs(folder_path)
        print(f"Папка {folder_path} была создана.")
    except Exception as e:
        print(f"Ошибка при создании папки: {e}")
else:
    print("Папка существует.")

file_path = os.path.join(folder_path, 'new_wb.xlsx')

try:
    wb.save(file_path)
    print(f"Файл успешно сохранен: {file_path}")
except Exception as e:
    print(f"Произошла ошибка при сохранении файла: {e}")
