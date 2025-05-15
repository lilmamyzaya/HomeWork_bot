import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

#Функция для форматирования столбцов по длине
def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 1

# Функция для извлечения даты из имени файла
def extract_date_from_filename(file_name):

    date_pattern = r"tasks_(\d{2}\.\d{2}\.\d{4})"
    match = re.search(date_pattern, file_name)
    if match:
        return match.group(1)
    return None

# Функция для записи даты в файл
def write_to_excel_data(date, excel_folder_path, output_file="new_wb.xlsx"):
    excel_file_path = os.path.join(excel_folder_path, output_file)

    if os.path.exists(excel_file_path):
        wb = load_workbook(excel_file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    next_column = ws.max_column + 1
    ws.cell(row=1, column=next_column, value=date)

    auto_adjust_column_width(ws)

    wb.save(excel_file_path)
    print(f"Дата {date} добавлена в файл: {excel_file_path}")


# Основная функция для обработки файлов в указанной папке
def process_files_data(folder_path_tasks, excel_folder_path):
    for file_name in os.listdir(folder_path_tasks):
        if file_name.startswith("tasks_"):
            date = extract_date_from_filename(file_name)
            if date:
                print(f"Дата извлечена: {date}")
                write_to_excel_data(date, excel_folder_path)
            else:
                print(f"В файле {file_name} не найдена дата.")

# Функция для извлечения фио из имени файла
def extract_surname_from_filename(file_name):
    """name_pattern = r"^(\S+)
    match = re.match(name_pattern, file_name)
    if match:
        return match.group(1)
    return None"""
    return os.path.splitext(file_name)[0]

def write_to_excel_name(surname, excel_folder_path, output_file="new_wb.xlsx"):
    excel_file_path = os.path.join(excel_folder_path, output_file)

    if os.path.exists(excel_file_path):
        wb = load_workbook(excel_file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    ws['A1'] = "Дата, на которую было дз"
    ws['A2'] = "Номера задач"      

    if ws.max_row == 1 and ws['A1'].value == "Дата, на которую было дз":
        next_row = 3  # Начинаем с третьей строки, после заголовков
    else:
        next_row = ws.max_row + 1

    ws[f'A{next_row}'] = surname

    auto_adjust_column_width(ws)

    wb.save(excel_file_path)
    print(f"Фамилия {surname} добавлена в файл: {excel_file_path}")

def process_files_surname(folder_path_students, excel_folder_path):
    for file_name in os.listdir(folder_path_students):
        surname = extract_surname_from_filename(file_name)
        if surname:
            print(f"Фамилия извлечена: {surname}")
            write_to_excel_name(surname, excel_folder_path)
        else:
            print(f"В файле {file_name} не найдена фамилия.")

folder_path_tasks = r"C:\Users\Инстаблогерша\OneDrive\Desktop\bot_py\hw_on_differential_equations_bot\tasks"  # Папка с файлами задач
excel_folder_path = r"C:\Users\Инстаблогерша\OneDrive\Desktop\bot_py\hw_on_differential_equations_bot\bot\data"  # Папка для Excel-файла
folder_path_students = r"C:\Users\Инстаблогерша\OneDrive\Desktop\bot_py\hw_on_differential_equations_bot\students" # Папка с файлами студентов

process_files_data(folder_path_tasks, excel_folder_path)
process_files_surname(folder_path_students, excel_folder_path)
